#!/usr/bin/env python3
"""
Parse une matrice pédagogique Excel et extrait le programme (UE + matières).

Usage: python scripts/parse_matrice.py <fichier_excel>
Exemple: python scripts/parse_matrice.py sources/matrices/matrice_achats.xlsx

Affiche le programme structuré en YAML pour copier-coller dans le fichier formation.
"""

import sys
import os
import json

try:
    import openpyxl
except ImportError:
    print("Installation des dépendances...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl


def parse_matrice(filepath: str) -> dict:
    """
    Parse une matrice pédagogique Excel.

    La structure attendue peut varier, mais en général :
    - Les onglets correspondent aux années
    - Les lignes contiennent les UE et matières
    - Les UE sont identifiables par leur mise en forme (gras, couleur de fond)

    Cette fonction tente une détection heuristique. Si le format ne correspond pas,
    elle affiche le contenu brut pour que l'utilisateur puisse guider le parsing.
    """

    wb = openpyxl.load_workbook(filepath, data_only=True)
    programme = {}

    print(f"Fichier: {filepath}")
    print(f"Onglets: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== Onglet: {sheet_name} ===")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Lignes: {ws.max_row}, Colonnes: {ws.max_column}")
        print()

        current_ue = None
        annee_data = {}

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column, 10)):
            # Récupérer les valeurs non vides
            values = []
            is_bold = False
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value).strip())
                    # Détecter si c'est une UE (souvent en gras ou avec un fond coloré)
                    if cell.font and cell.font.bold:
                        is_bold = True

            if not values:
                continue

            line = " | ".join(values)

            # Heuristique: si le texte est en gras ou contient des mots-clés d'UE
            ue_keywords = ["UE", "Bloc", "Module", "Unité", "BLOC", "MODULE"]
            is_ue = is_bold or any(kw in values[0] for kw in ue_keywords)

            if is_ue and len(values[0]) > 3:
                current_ue = values[0]
                if current_ue not in annee_data:
                    annee_data[current_ue] = []
                print(f"  [UE] {current_ue}")
            elif current_ue and values:
                matiere = values[0]
                annee_data[current_ue].append(matiere)
                print(f"        - {matiere}")
            else:
                print(f"  ??? {line}")

        if annee_data:
            programme[sheet_name] = annee_data
        print()

    # Afficher en format YAML pour copier-coller
    print("\n" + "=" * 60)
    print("PROGRAMME AU FORMAT YAML (à copier dans le fichier formation)")
    print("=" * 60)
    print("programme:")
    for annee, ues in programme.items():
        print(f'  "{annee}":')
        for ue, matieres in ues.items():
            print(f'    "{ue}":')
            for m in matieres:
                print(f'      - "{m}"')

    return programme


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/parse_matrice.py <fichier_excel>")
        print("Le script affiche le programme structuré pour le copier dans le YAML.")
        sys.exit(1)

    parse_matrice(sys.argv[1])
